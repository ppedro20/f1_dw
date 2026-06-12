import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Color palette ──
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
LABEL_FILL = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FF1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="FF2E75B6")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
MATCH_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_label_cell(ws, row, col, text=None):
    cell = ws.cell(row=row, column=col)
    cell.fill = LABEL_FILL
    cell.font = LABEL_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border
    if text:
        cell.value = text
    return cell

def style_body_cell(ws, row, col, value=None, wrap=True):
    cell = ws.cell(row=row, column=col)
    cell.font = BODY_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = thin_border
    if value:
        cell.value = value
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def merge_and_style(ws, range_str, value=None, font=None, fill=None, alignment=None, border=None):
    ws.merge_cells(range_str)
    top_left = range_str.split(":")[0]
    cell = ws[top_left]
    if value:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

# ══════════════════════════════════════════════════════════
# SHEET 1: Matriz
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Matriz"
set_col_widths(ws1, [22, 28, 18, 18, 18, 18, 18, 18, 18, 22])

merge_and_style(ws1, "A1:J1", "Matriz Facto-Dimensão — Modelo Dimensional F1", TITLE_FONT)
merge_and_style(ws1, "A2:J2",
    "Assunto: Performance em Corrida | Granularidade: 1 linha por piloto por corrida",
    Font(name="Calibri", italic=True, size=10, color="FF666666"))

# Header row
headers = ["Assunto / Factos", "Medidas (Factos)", "Dim_Tempo", "Dim_Piloto", "Dim_Circuito",
           "Dim_Construtor", "Dimensão 6", "Dimensão 7", "Dimensão 8", "Contagem"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 10)

# ASSUNTO-1: Performance em Corrida (rows 5-9)
merge_and_style(ws1, "A5:A9", "Performance em Corrida")
style_label_cell(ws1, 5, 1)
for r in range(6, 10):
    style_label_cell(ws1, r, 1)

facts = [
    "Pontos_Conquistados",
    "Tempo_Total_Pit_Stops",
    "Posicoes_Ganhas",
    "Posicao_Partida",
    "Posicao_Final",
]

for i, fname in enumerate(facts):
    row = 5 + i
    style_body_cell(ws1, row, 2, fname)

# Relationship matrix (1 = related) — all 5 facts relate to all 4 dimensions
for i in range(5):
    row = 5 + i
    for col in range(3, 7):  # Dimensões 1-4
        c = ws1.cell(row=row, column=col, value=1)
        c.fill = MATCH_FILL
        c.font = Font(name="Calibri", size=10, bold=True, color="FF006100")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

# Count formulas per row
for i in range(5):
    row = 5 + i
    ws1.cell(row=row, column=10, value=f'=COUNTIF(C{row}:I{row},"1")')
    style_body_cell(ws1, row, 10)

# Bottom counts row
count_row = 11
style_label_cell(ws1, count_row, 2, "Contagem: Utilização dos Factos")
for col in range(3, 10):
    lbl = get_column_letter(col)
    c = ws1.cell(row=count_row, column=col, value=f'=COUNTIF({lbl}5:{lbl}9,1)')
    c.font = Font(name="Calibri", bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

# Legend
merge_and_style(ws1, "A13:J13",
    "Legenda: Coloque o número 1 para indicar que existe relacionamento entre a dimensão e a medida.",
    Font(name="Calibri", italic=True, size=9, color="FF666666"))

# ══════════════════════════════════════════════════════════
# DIMENSION SHEET BUILDER
# ══════════════════════════════════════════════════════════
def build_dimension_sheet(wb, sheet_name, dim_name, description, source_table,
                          attributes, obs, sql_query):
    ws = wb.create_sheet(title=sheet_name)
    set_col_widths(ws, [18, 22, 28, 10, 20, 22, 22, 28])

    # Title
    merge_and_style(ws, "A1:H1", f"Ficha de Dimensão — {dim_name}", TITLE_FONT)

    # Row 3: Dimension name
    style_label_cell(ws, 3, 1, "Nome da Dimensão:")
    ws["B3"] = dim_name
    ws["B3"].fill = YELLOW_FILL
    ws["B3"].font = Font(name="Calibri", bold=True, size=12, color="FF1F4E79")

    # Row 5: Description label (A5) | Source label (E5)
    style_label_cell(ws, 5, 1, "Descrição")
    style_label_cell(ws, 5, 5, "Tabela/Atributo Original")

    # Row 6: Description value (merged B6:D6) | Source value (merged F6:H6)
    merge_and_style(ws, "B6:D6", description, BODY_FONT, alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(2, 5):
        ws.cell(row=6, column=c).border = thin_border

    merge_and_style(ws, "F6:H6", source_table, BODY_FONT, alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(6, 9):
        ws.cell(row=6, column=c).border = thin_border

    # Borders around the label rows
    for col in range(1, 5):
        ws.cell(row=5, column=col).border = thin_border
        ws.cell(row=6, column=col).border = thin_border
    for col in range(5, 9):
        ws.cell(row=5, column=col).border = thin_border
        ws.cell(row=6, column=col).border = thin_border

    # Row 7: Attributes header
    attr_headers = ["Atributos Importantes", "Nome", "Descrição", "Muda?",
                    "Tabela", "Atributo", "Relacionamento", "Regras/Domínio"]
    for i, h in enumerate(attr_headers, 1):
        ws.cell(row=8, column=i, value=h)
    style_header_row(ws, 8, 8)

    # Attribute rows (starting at row 9)
    for i, (nome, desc, muda, tabela, atributo, relac, regras) in enumerate(attributes):
        row = 9 + i
        if i == 0:
            style_label_cell(ws, row, 1, "Atributos")
        style_body_cell(ws, row, 2, nome)
        style_body_cell(ws, row, 3, desc)
        c = style_body_cell(ws, row, 4, muda)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if muda and ("Sim" in str(muda) or "SCD" in str(muda)):
            c.font = Font(name="Calibri", size=10, bold=True, color="FFC00000")
        style_body_cell(ws, row, 5, tabela)
        style_body_cell(ws, row, 6, atributo)
        style_body_cell(ws, row, 7, relac)
        style_body_cell(ws, row, 8, regras)

    # Observations
    obs_row = 9 + len(attributes) + 1
    style_label_cell(ws, obs_row, 1, "Obs.:")
    merge_and_style(ws, f"B{obs_row}:H{obs_row}", obs,
                    Font(name="Calibri", italic=True, size=10),
                    alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(2, 9):
        ws.cell(row=obs_row, column=c).border = thin_border

    # SQL section
    sql_hdr = obs_row + 2
    merge_and_style(ws, f"A{sql_hdr}:H{sql_hdr}",
        "Como obter a informação do sistema origem (incluir SQL se necessário)",
        Font(name="Calibri", bold=True, size=10, color="FF1F4E79"))
    for c in range(1, 9):
        ws.cell(row=sql_hdr, column=c).border = thin_border

    sql_row = sql_hdr + 1
    merge_and_style(ws, f"A{sql_row}:H{sql_row}", sql_query,
                    Font(name="Consolas", size=9),
                    alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(1, 9):
        ws.cell(row=sql_row, column=c).border = thin_border

    return ws

# ══════════════════════════════════════════════════════════
# FACT SHEET BUILDER
# ══════════════════════════════════════════════════════════
def build_fact_sheet(wb, sheet_name, fact_name, description, data_origin,
                     measures, relationships, obs, sql_query):
    ws = wb.create_sheet(title=sheet_name)
    set_col_widths(ws, [18, 22, 28, 14, 14, 20, 22, 22, 28])

    # Title
    merge_and_style(ws, "A1:I1", f"Ficha de Facto — {fact_name}", TITLE_FONT)

    # Row 3: Fact name
    style_label_cell(ws, 3, 1, "Nome:")
    ws["B3"] = fact_name
    ws["B3"].fill = YELLOW_FILL
    ws["B3"].font = Font(name="Calibri", bold=True, size=12, color="FF1F4E79")

    # Row 5: Description label (A5) | Data origin label (F5)
    style_label_cell(ws, 5, 1, "Descrição")
    style_label_cell(ws, 5, 6, "Origem dos dados")

    # Row 6: Description value (merged B6:E6) | Data origin value (merged G6:I6)
    merge_and_style(ws, "B6:E6", description, BODY_FONT, alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(2, 6):
        ws.cell(row=6, column=c).border = thin_border

    merge_and_style(ws, "G6:I6", data_origin, BODY_FONT, alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(7, 10):
        ws.cell(row=6, column=c).border = thin_border

    # Borders
    for col in range(1, 6):
        ws.cell(row=5, column=col).border = thin_border
        ws.cell(row=6, column=col).border = thin_border
    for col in range(6, 10):
        ws.cell(row=5, column=col).border = thin_border
        ws.cell(row=6, column=col).border = thin_border

    # Row 8: Measures header
    meas_headers = ["Medidas", "Nome", "Descrição", "Formato", "Calculado?",
                    "Tabela", "Atributo", "Relacionamento", "Regras/Domínio"]
    for i, h in enumerate(meas_headers, 1):
        ws.cell(row=8, column=i, value=h)
    style_header_row(ws, 8, 9)

    # Measure rows (starting at row 9)
    for i, (nome, desc, formato, calc, tabela, atributo, relac, regras) in enumerate(measures):
        row = 9 + i
        if i == 0:
            style_label_cell(ws, row, 1, "Medidas")
        style_body_cell(ws, row, 2, nome)
        style_body_cell(ws, row, 3, desc)
        style_body_cell(ws, row, 4, formato)
        c = style_body_cell(ws, row, 5, calc)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if calc and calc.strip().lower().startswith("sim"):
            c.font = Font(name="Calibri", size=10, italic=True, color="FFC00000")
        style_body_cell(ws, row, 6, tabela)
        style_body_cell(ws, row, 7, atributo)
        style_body_cell(ws, row, 8, relac)
        style_body_cell(ws, row, 9, regras)

    # Relationships section
    rel_start = 9 + len(measures) + 1
    ws.cell(row=rel_start, column=1, value="Relacionamentos")
    ws.cell(row=rel_start, column=1).font = Font(name="Calibri", bold=True, size=11, color="FF1F4E79")
    for c in range(1, 10):
        ws.cell(row=rel_start, column=c).border = thin_border

    rel_headers = ["", "Nome", "Ligação Facto", "Ligação Dimensão"]
    hdr_row = rel_start + 1
    for i, h in enumerate(rel_headers, 1):
        if h:
            ws.cell(row=hdr_row, column=i, value=h)
    style_header_row(ws, hdr_row, 9)

    for i, (nome, lig_facto, lig_dim) in enumerate(relationships):
        row = hdr_row + 1 + i
        style_body_cell(ws, row, 2, nome)
        style_body_cell(ws, row, 3, lig_facto)
        style_body_cell(ws, row, 4, lig_dim)
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = thin_border

    # Observations
    obs_row = hdr_row + 1 + len(relationships) + 1
    style_label_cell(ws, obs_row, 1, "Obs.:")
    merge_and_style(ws, f"B{obs_row}:I{obs_row}", obs,
                    Font(name="Calibri", italic=True, size=10),
                    alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(2, 10):
        ws.cell(row=obs_row, column=c).border = thin_border

    # SQL section
    sql_hdr = obs_row + 2
    merge_and_style(ws, f"A{sql_hdr}:G{sql_hdr}",
        "Como obter a informação do sistema origem (incluir SQL se necessário)",
        Font(name="Calibri", bold=True, size=10, color="FF1F4E79"))
    for c in range(1, 9):
        ws.cell(row=sql_hdr, column=c).border = thin_border

    sql_row = sql_hdr + 1
    merge_and_style(ws, f"A{sql_row}:G{sql_row}", sql_query,
                    Font(name="Consolas", size=9),
                    alignment=Alignment(wrap_text=True, vertical="top"))
    for c in range(1, 9):
        ws.cell(row=sql_row, column=c).border = thin_border

    return ws

# ══════════════════════════════════════════════════════════
# BUILD ALL SHEETS
# ══════════════════════════════════════════════════════════

# Dim_Tempo
build_dimension_sheet(wb, "Dim_Tempo", "Dim_Tempo",
    "Dimensão temporal com granularidade ao nível do dia. Hierarquia: Ano → Mês → Dia.",
    "races.date → DATE → YYYYMMDD",
    [
        ("Data_SK", "Chave surrogate (YYYYMMDD)", "Não", "races", "date",
         "Chave Primária", "DATE → INTEGER YYYYMMDD"),
        ("Ano", "Ano da corrida", "Não", "races", "date",
         "Derivado", "YEAR(date)"),
        ("Mes", "Mês da corrida (1-12)", "Não", "races", "date",
         "Derivado", "MONTH(date)"),
        ("Dia", "Dia da corrida (1-31)", "Não", "races", "date",
         "Derivado", "DAY(date)"),
    ],
    "SCD Tipo 0 — atributos estáticos, sem histórico.",
    "SELECT DISTINCT\n"
    "  CAST(CONVERT(VARCHAR, date, 112) AS INT) AS Data_SK,\n"
    "  YEAR(date) AS Ano,\n"
    "  MONTH(date) AS Mes,\n"
    "  DAY(date) AS Dia\n"
    "FROM races\n"
    "ORDER BY date;"
)

# Dim_Piloto
build_dimension_sheet(wb, "Dim_Piloto", "Dim_Piloto",
    "Dimensão de pilotos. SCD Tipo 2 no atributo Equipa_Atual para preservar histórico de mudanças de equipa.",
    "drivers + results + constructors",
    [
        ("Piloto_SK", "Chave surrogate", "Não", "drivers", "driverId",
         "Chave Primária", "Identity / sequencial"),
        ("Nome_Completo", "Nome completo do piloto", "Não", "drivers", "forename, surname",
         "Concatenação", "CONCAT(forename, ' ', surname)"),
        ("Nacionalidade", "País de nacionalidade", "Não", "drivers", "nationality",
         "Direto", "Mapeamento direto"),
        ("Data_Nascimento", "Data de nascimento", "Não", "drivers", "dob",
         "Direto", "NULL se não disponível"),
        ("Equipa_Atual", "Equipa do piloto na época/corrida", "Sim (SCD Tipo 2)", "constructors", "name",
         "FK via results.constructorId", "JOIN results → constructors; nova linha se equipa mudar"),
    ],
    "SCD Tipo 2 em Equipa_Atual. Demais atributos SCD Tipo 0. "
    "Data de vigência (StartDate/EndDate) e flag CurrentRecord para rastrear histórico.",
    "SELECT DISTINCT\n"
    "  d.driverId,\n"
    "  CONCAT(d.forename, ' ', d.surname) AS Nome_Completo,\n"
    "  d.nationality AS Nacionalidade,\n"
    "  d.dob AS Data_Nascimento,\n"
    "  c.name AS Equipa_Atual\n"
    "FROM drivers d\n"
    "JOIN results r ON d.driverId = r.driverId\n"
    "JOIN constructors c ON r.constructorId = c.constructorId\n"
    "ORDER BY d.driverId;"
)

# Dim_Circuito
build_dimension_sheet(wb, "Dim_Circuito", "Dim_Circuito",
    "Dimensão de circuitos com hierarquia: Continente → País → Cidade → Circuito.",
    "circuits",
    [
        ("Circuito_SK", "Chave surrogate", "Não", "circuits", "circuitId",
         "Chave Primária", "Identity / sequencial"),
        ("Nome_Circuito", "Nome oficial do circuito", "Não", "circuits", "name",
         "Direto", "Mapeamento direto"),
        ("Cidade", "Cidade de localização", "Não", "circuits", "location",
         "Direto", "NULL se não disponível"),
        ("Pais", "País do circuito", "Não", "circuits", "country",
         "Direto", "Normalizar 'USA' → 'United States'"),
        ("Continente", "Continente do circuito", "Não", "circuits", "country",
         "Lookup geográfica", "Tabela de lookup país → continente"),
    ],
    "SCD Tipo 0. Hierarquia: Continente → País → Cidade → Circuito. "
    "Nota: circuits.country tem inconsistências (ex: 'USA' vs 'United States') — normalizar na ETL.",
    "SELECT\n"
    "  circuitId AS Circuito_SK,\n"
    "  name AS Nome_Circuito,\n"
    "  location AS Cidade,\n"
    "  country AS Pais\n"
    "  -- Continente: lookup externa por país\n"
    "FROM circuits\n"
    "ORDER BY country, location;"
)

# Dim_Construtor
build_dimension_sheet(wb, "Dim_Construtor", "Dim_Construtor",
    "Dimensão de construtores/equipas. SCD Tipo 1 no atributo Motorizador.",
    "constructors",
    [
        ("Construtor_SK", "Chave surrogate", "Não", "constructors", "constructorId",
         "Chave Primária", "Identity / sequencial"),
        ("Nome", "Nome oficial da equipa", "Não", "constructors", "name",
         "Direto", "Mapeamento direto"),
        ("Pais", "País de origem da equipa", "Não", "constructors", "nationality",
         "Direto", "Mapeamento direto"),
        ("Motorizador", "Fornecedor de motor", "Sim (SCD Tipo 1)", "constructors", "name",
         "Lookup", "Regras: Mercedes-AMG → Mercedes, RBPT → Red Bull, Ferrari → Ferrari"),
    ],
    "SCD Tipo 1 para Motorizador (sobrescrita). Demais atributos SCD Tipo 0.",
    "SELECT\n"
    "  constructorId AS Construtor_SK,\n"
    "  name AS Nome,\n"
    "  nationality AS Pais\n"
    "  -- Motorizador: lookup com regras de negócio\n"
    "FROM constructors\n"
    "ORDER BY name;"
)

# Fact_Performance
build_fact_sheet(wb, "Fact_Performance", "Fact_Performance",
    "Fact table de performance em corrida. Granularidade: 1 registo por piloto por corrida. "
    "Mede pontos conquistados, tempo total em pit stops, posições ganhas/perdidas, "
    "posição de partida e posição final.",
    "results (principal) + pit_stops (agregação) + races (data FK)",
    [
        ("Pontos_Conquistados", "Pontos obtidos pelo piloto na corrida",
         "DECIMAL(8,2)", "Não", "results", "points",
         "Direto", "≥ 0; sistema de pontuação F1"),
        ("Tempo_Total_Pit_Stops", "Soma da duração de todas as paragens do piloto",
         "DECIMAL(8,3)", "Sim (agregação)", "pit_stops", "duration",
         "SUM por raceId + driverId", "Agrupar todas as paragens do piloto na corrida"),
        ("Posicoes_Ganhas", "Diferença entre posição de partida e final",
         "INTEGER", "Sim (derivado)", "results", "grid, position",
         "Cálculo: grid - position", "0 se position IS NULL"),
        ("Posicao_Partida", "Posição na grelha de partida",
         "INTEGER", "Não", "results", "grid",
         "Direto", "≥ 1"),
        ("Posicao_Final", "Posição final na corrida",
         "INTEGER", "Não", "results", "position",
         "Direto", "NULL = não classificado"),
    ],
    [
        ("FK_Tempo", "Data_SK", "Dim_Tempo.Data_SK"),
        ("FK_Piloto", "Piloto_SK", "Dim_Piloto.Piloto_SK"),
        ("FK_Circuito", "Circuito_SK", "Dim_Circuito.Circuito_SK"),
        ("FK_Construtor", "Construtor_SK", "Dim_Construtor.Construtor_SK"),
    ],
    "Granularidade: 1 linha por (piloto, corrida). Refrescamento incremental por data "
    "(races.date). Tratar NULLs em position como 'Não classificado'.",
    "SELECT\n"
    "  CAST(CONVERT(VARCHAR, r2.date, 112) AS INT) AS Data_SK,\n"
    "  d.driverId AS Piloto_SK,\n"
    "  c2.circuitId AS Circuito_SK,\n"
    "  c.constructorId AS Construtor_SK,\n"
    "  r.points AS Pontos_Conquistados,\n"
    "  COALESCE(ps.total_pit_duration, 0) AS Tempo_Total_Pit_Stops,\n"
    "  COALESCE(r.grid - r.position, 0) AS Posicoes_Ganhas,\n"
    "  r.grid AS Posicao_Partida,\n"
    "  r.position AS Posicao_Final\n"
    "FROM results r\n"
    "JOIN races r2 ON r.raceId = r2.raceId\n"
    "JOIN drivers d ON r.driverId = d.driverId\n"
    "JOIN constructors c ON r.constructorId = c.constructorId\n"
    "JOIN circuits c2 ON r2.circuitId = c2.circuitId\n"
    "LEFT JOIN (\n"
    "  SELECT raceId, driverId, SUM(duration) AS total_pit_duration\n"
    "  FROM pit_stops GROUP BY raceId, driverId\n"
    ") ps ON r.raceId = ps.raceId AND r.driverId = ps.driverId\n"
    "ORDER BY r2.date, r.positionOrder;"
)

# ══════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════
output_path = r"C:\dev\Projects\f1_dw\project\3_modelo_dimensional\F1_Data_Warehouse_Modelo_Dimensional.xlsx"
wb.save(output_path)
print(f"OK - File saved to: {output_path}")
print(f"  Sheets: {wb.sheetnames}")
